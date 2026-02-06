"""
Completed + Upcoming Revenue → SharePoint Excel Updater (Outlook Graph + openpyxl)
===============================================================================

Replaces:
- Gmail IMAP + gspread (Google Sheets)
With:
- Microsoft Graph (Outlook mailbox) + SharePoint Excel file (download/edit/upload)

Logic preserved from your script:
1) Upcoming: sum(Subtotal) by date for dates AFTER "today" date found in attachment,
   overwrite Scheduled Revenue columns (global + BU scheduled).
2) Completed: SUM all Jobs Subtotal values (ignore date) per your updated rule,
   set Completed Revenue for file_date, clear Scheduled Revenue for that date
   (global + BU scheduled cleared).

Email subjects (contain-match via Graph $search):
- "Completed Revenue for Retail Excel Dashboard"
- "Upcoming Revenue for Retail Excel Dashboard"

SharePoint Excel sheet:
- TAB_NAME default = "February"
- Base columns:
  Col2 Date (match key; do NOT update)
  Col4 Completed Revenue
  Col5 Scheduled Revenue
- BU_MAP controls BU date/completed/scheduled columns

Required env vars:
- TENANT_ID
- CLIENT_ID
- CLIENT_SECRET
- MAILBOX_UPN   (e.g., apatil@glassdoctordfw.com)
- DRIVE_ID      (SharePoint drive id)
- FILE_ITEM_ID  (Excel file item id)

Install:
pip install msal requests openpyxl
"""

import os, io, re, base64
from datetime import datetime, date, timezone
from typing import Optional, Dict, Any, List, Tuple

import requests
from msal import ConfidentialClientApplication
from openpyxl import load_workbook


# -------------------- CONFIG --------------------
TAB_NAME = os.getenv("TAB_NAME", "February")

SUBJECT_COMPLETED_PHRASE = "Completed Revenue for Retail Excel Dashboard"
SUBJECT_UPCOMING_PHRASE  = "Upcoming Revenue for Retail Excel Dashboard"

# Base sheet columns (1-based)
COL_DATE      = 2   # B
COL_COMPLETED = 4   # D
COL_SCHEDULED = 5   # E

# -------------------- BU column mappings --------------------
BU_MAP = {
    "arlington":   {"date_col": 9,  "completed_col": 11, "scheduled_col": 12},  # I, K, L
    "carrollton":  {"date_col": 16, "completed_col": 18, "scheduled_col": 19},  # P, R, S
    "colleyville": {"date_col": 23, "completed_col": 25, "scheduled_col": 26},  # W, Y, Z
    "dallas":      {"date_col": 30, "completed_col": 32, "scheduled_col": 33},  # AD, AF, AG
    "denton":      {"date_col": 37, "completed_col": 39, "scheduled_col": 40},  # AK, AM, AN
}

GRAPH = "https://graph.microsoft.com/v1.0"


# -------------------- Small helpers --------------------
def must_env(name: str) -> str:
    v = os.getenv(name)
    if not v:
        raise RuntimeError(f"Missing environment variable: {name}")
    return v

def parse_dt(dt_str: str) -> datetime:
    if not dt_str:
        return datetime(1970, 1, 1, tzinfo=timezone.utc)
    if dt_str.endswith("Z"):
        dt_str = dt_str.replace("Z", "+00:00")
    return datetime.fromisoformat(dt_str)

def col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def parse_money(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r"[^0-9.\-]", "", str(val))
    return float(s) if s else 0.0

def try_parse_any_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def find_col_idx(header, target_names_lower):
    h = [str(x).strip().lower() for x in header]
    for i, name in enumerate(h):
        if name in target_names_lower:
            return i
    for i, name in enumerate(h):
        for t in target_names_lower:
            if t in name:
                return i
    return None


# -------------------- Auth (App-only) --------------------
def get_token() -> str:
    tenant_id = os.getenv("tenant_id")
    client_id = os.getenv("client_id")
    client_secret = os.getenv("client_secret")
    
    app = ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Token error: {result.get('error')} - {result.get('error_description')}")
    return result["access_token"]

def graph_get(token: str, url: str, params: Optional[dict] = None) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",  # required for $search
    }
    r = requests.get(url, headers=headers, params=params, timeout=60)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.json()

def graph_get_bytes(token: str, url: str, params: Optional[dict] = None) -> bytes:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }
    r = requests.get(url, headers=headers, params=params, timeout=120)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.content

def graph_put_bytes(token: str, url: str, content: bytes, content_type: str) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": content_type,
    }
    r = requests.put(url, headers=headers, data=content, timeout=180)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.json()


# -------------------- Outlook: latest email + attachment --------------------
def latest_message_for_subject(token: str, mailbox_upn: str, subject_phrase: str) -> Optional[Dict[str, Any]]:
    """
    Uses $search (contains-style) and sorts locally (since $search can't combine with $orderby).
    """
    url = f"{GRAPH}/users/{mailbox_upn}/mailFolders/Inbox/messages"
    params = {
        "$select": "id,subject,receivedDateTime,from,hasAttachments",
        "$top": "25",
        "$search": f"\"{subject_phrase}\"",
    }
    data = graph_get(token, url, params=params)
    msgs: List[Dict[str, Any]] = data.get("value", [])

    # Prefer subject contains phrase (case-insensitive)
    phrase = subject_phrase.lower()
    candidates = [m for m in msgs if phrase in (m.get("subject") or "").lower()]

    if not candidates:
        return None

    candidates.sort(key=lambda m: parse_dt(m.get("receivedDateTime", "")), reverse=True)
    return candidates[0]

def get_first_xlsx_attachment_from_message(token: str, mailbox_upn: str, message_id: str) -> Tuple[Optional[str], Optional[bytes]]:
    """
    Returns (filename, bytes) for the first .xlsx attachment on the message.
    Handles:
    - contentBytes if present
    - fallback to /$value download
    """
    # list attachments
    url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments"
    data = graph_get(token, url, params={"$top": "50"})
    atts = data.get("value", [])

    # pick first .xlsx
    for a in atts:
        name = (a.get("name") or "")
        if name.lower().endswith(".xlsx"):
            # small attachments often include contentBytes
            cb = a.get("contentBytes")
            if cb:
                return name, base64.b64decode(cb)

            # else download raw
            att_id = a.get("id")
            if att_id:
                raw_url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments/{att_id}/$value"
                b = graph_get_bytes(token, raw_url)
                return name, b

    return None, None


# -------------------- SharePoint Excel: download/edit/upload --------------------
def download_sharepoint_excel(token: str, drive_id: str, item_id: str) -> bytes:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    return graph_get_bytes(token, url)

def upload_sharepoint_excel(token: str, drive_id: str, item_id: str, content: bytes) -> None:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    graph_put_bytes(
        token,
        url,
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -------------------- Attachment parsing (kept style) --------------------
def read_xlsx_first_sheet_rows(xlsx_bytes: bytes) -> List[List[Any]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else v) for v in r])
    return rows

def extract_date_from_filename(fname: str) -> date:
    if not fname:
        raise ValueError("Missing filename for date extraction.")
    m = re.search(r"(\d{1,4})[._-](\d{1,2})[._-](\d{1,4})", fname)
    if not m:
        raise ValueError(f"Could not extract date from filename: {fname}")

    a, b, c = m.groups()
    nums = list(map(int, (a, b, c)))

    if nums[0] > 31:  # YYYY first
        yyyy, mm, dd = nums[0], nums[1], nums[2]
    else:  # MM first
        mm, dd, yy = nums
        yyyy = 2000 + yy if yy < 100 else yy

    return date(yyyy, mm, dd)

def subtotal_by_date_from_rows_upcoming(rows):
    if not rows or len(rows) < 2:
        raise ValueError("Attachment sheet is empty or missing header/body.")

    header = rows[0]
    body = rows[1:]

    bu_col   = find_col_idx(header, {"business unit"})
    date_col = find_col_idx(header, {"next appt start date"})
    sub_col  = find_col_idx(header, {"jobs subtotal", "subtotal"})

    if bu_col is None or date_col is None or sub_col is None:
        raise ValueError(f"Could not find required columns in Upcoming. Header: {header}")

    totals_by_bu_date = {bu: {} for bu in BU_MAP.keys()}
    dates_seen = []

    for r in body:
        if len(r) <= max(bu_col, date_col, sub_col):
            continue

        bu_raw = str(r[bu_col]).strip().lower()
        if bu_raw not in totals_by_bu_date:
            continue

        d = try_parse_any_date(r[date_col])
        if not d:
            continue

        subtotal = parse_money(r[sub_col])
        m = totals_by_bu_date[bu_raw]
        m[d] = m.get(d, 0.0) + subtotal
        dates_seen.append(d)

    if not dates_seen:
        raise ValueError("No valid dates found in Upcoming attachment.")

    today_in_file = min(dates_seen)
    for bu in totals_by_bu_date:
        totals_by_bu_date[bu] = {k: round(v, 2) for k, v in totals_by_bu_date[bu].items()}

    return today_in_file, totals_by_bu_date

def completed_values_from_rows_by_bu_sum_jobs_subtotal(rows):
    if not rows or len(rows) < 2:
        raise ValueError("Attachment sheet is empty or missing header/body.")

    header = rows[0]
    body = rows[1:]

    bu_col  = find_col_idx(header, {"business unit"})
    sub_col = find_col_idx(header, {"jobs subtotal", "subtotal"})

    if bu_col is None or sub_col is None:
        raise ValueError(f"Could not find required columns in Completed. Header: {header}")

    global_sum = 0.0
    bu_sum = {bu: 0.0 for bu in BU_MAP.keys()}

    for r in body:
        if len(r) <= max(bu_col, sub_col):
            continue

        v = r[sub_col]
        if v in (None, "", " "):
            continue

        amt = parse_money(v)
        global_sum += amt

        bu_raw = str(r[bu_col]).strip().lower()
        if bu_raw in bu_sum:
            bu_sum[bu_raw] += amt

    global_sum = round(global_sum, 2)
    bu_sum = {bu: round(val, 2) for bu, val in bu_sum.items()}

    return global_sum, bu_sum


# -------------------- SharePoint workbook updates --------------------
def build_sheet_date_row_map_xl(ws, date_col: int) -> Dict[date, int]:
    """
    Reads a date column and returns {date: row_number}
    Assumes row 1 is header.
    """
    mapping = {}
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        v = ws.cell(row=row, column=date_col).value
        d = try_parse_any_date(v)
        if d:
            mapping[d] = row
    return mapping

def apply_upcoming_to_workbook(wb, sheet_name: str, today_in_file: date, totals_by_bu_date: Dict[str, Dict[date, float]]) -> int:
    ws = wb[sheet_name]

    base_date_row = build_sheet_date_row_map_xl(ws, COL_DATE)
    bu_date_row = {bu: build_sheet_date_row_map_xl(ws, BU_MAP[bu]["date_col"]) for bu in BU_MAP}

    # global totals = sum across BU totals by date
    global_totals: Dict[date, float] = {}
    for bu, dmap in totals_by_bu_date.items():
        for d, amt in dmap.items():
            global_totals[d] = global_totals.get(d, 0.0) + amt
    global_totals = {d: round(v, 2) for d, v in global_totals.items()}

    updated_cells = 0

    # Global scheduled updates (Col5) for d > today_in_file
    for d, total in global_totals.items():
        if d <= today_in_file:
            continue
        row = base_date_row.get(d)
        if not row:
            continue
        ws.cell(row=row, column=COL_SCHEDULED).value = total
        updated_cells += 1

    # BU scheduled updates
    for bu, dmap in totals_by_bu_date.items():
        sched_col = BU_MAP[bu]["scheduled_col"]
        row_map = bu_date_row[bu]
        for d, total in dmap.items():
            if d <= today_in_file:
                continue
            row = row_map.get(d)
            if not row:
                continue
            ws.cell(row=row, column=sched_col).value = total
            updated_cells += 1

    return updated_cells

def apply_completed_to_workbook(wb, sheet_name: str, file_date: date, global_completed_value: float, bu_completed_values: Dict[str, float]) -> int:
    ws = wb[sheet_name]

    updates = 0

    # Base row map
    base_date_row = build_sheet_date_row_map_xl(ws, COL_DATE)
    base_row = base_date_row.get(file_date)
    if base_row:
        ws.cell(row=base_row, column=COL_COMPLETED).value = global_completed_value
        ws.cell(row=base_row, column=COL_SCHEDULED).value = ""  # clear scheduled
        updates += 2

    # BU-specific rows
    for bu, cols in BU_MAP.items():
        row_map = build_sheet_date_row_map_xl(ws, cols["date_col"])
        row = row_map.get(file_date)
        if not row:
            continue
        ws.cell(row=row, column=cols["completed_col"]).value = bu_completed_values.get(bu, 0.0)
        ws.cell(row=row, column=cols["scheduled_col"]).value = ""  # clear scheduled
        updates += 2

    return updates


# -------------------- Main --------------------
def main():
    token = get_token()

    mailbox_upn = "apatil@glassdoctordfw.com"  # apatil@glassdoctordfw.com
    drive_id = os.getenv("drive_id")
    file_item_id = os.getenv("file_item_id")

    # 1) Get latest messages
    up_msg = latest_message_for_subject(token, mailbox_upn, SUBJECT_UPCOMING_PHRASE)
    c_msg  = latest_message_for_subject(token, mailbox_upn, SUBJECT_COMPLETED_PHRASE)

    if not up_msg and not c_msg:
        print("No matching emails found for Upcoming or Completed.")
        return

    # 2) Download the SharePoint workbook once
    xls_bytes = download_sharepoint_excel(token, drive_id, file_item_id)
    wb = load_workbook(io.BytesIO(xls_bytes))

    if TAB_NAME not in wb.sheetnames:
        raise RuntimeError(f"Tab '{TAB_NAME}' not found in workbook. Found: {wb.sheetnames}")

    # 3) Upcoming: attachment -> parse -> apply
    if up_msg:
        up_id = up_msg["id"]
        up_fname, up_content = get_first_xlsx_attachment_from_message(token, mailbox_upn, up_id)
        if up_content:
            up_rows = read_xlsx_first_sheet_rows(up_content)
            up_today, totals_by_bu_date = subtotal_by_date_from_rows_upcoming(up_rows)
            n = apply_upcoming_to_workbook(wb, TAB_NAME, up_today, totals_by_bu_date)
            print(f"[Upcoming] {up_fname} | file_today={up_today} | updated cells={n}")
        else:
            print("[Upcoming] Email found but no .xlsx attachment.")
    else:
        print("[Upcoming] No matching email found.")

    # 4) Completed: attachment -> parse -> apply
    if c_msg:
        c_id = c_msg["id"]
        c_fname, c_content = get_first_xlsx_attachment_from_message(token, mailbox_upn, c_id)
        if c_content:
            c_rows = read_xlsx_first_sheet_rows(c_content)
            file_date = extract_date_from_filename(c_fname)

            global_sum, bu_sum = completed_values_from_rows_by_bu_sum_jobs_subtotal(c_rows)
            n = apply_completed_to_workbook(wb, TAB_NAME, file_date, global_sum, bu_sum)

            print(f"[Completed] {c_fname} | file_date={file_date} | global_completed={global_sum} | updated cells={n}")
        else:
            print("[Completed] Email found but no .xlsx attachment.")
    else:
        print("[Completed] No matching email found.")

    # 5) Save workbook back to bytes and upload
    out = io.BytesIO()
    wb.save(out)
    upload_sharepoint_excel(token, drive_id, file_item_id, out.getvalue())

    print("Done. Uploaded updated Excel back to SharePoint.")


if __name__ == "__main__":
    main()
